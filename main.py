from openpyxl import load_workbook
from openpyxl.styles import Font
import time

input('''Make sure that you have input the puzzle in
the data.xlsx excel file. If done, close the file
and continue[press enter]>>''')
print('Starting..')

def assign_values_to_vars(box_code, value):
    data[box_code] = value
    globals()[box_code[0]][box_code] = value
    globals()[box_code[1]][box_code] = value


def write_qnum_in_sheet(box_code, value):
    ws[box_code_to_sheet_code(box_code)].font = fontq
    ws[box_code_to_sheet_code(box_code)] = value
    wb.save("Data.xlsx")


def write_num_in_sheet(box_code, value):
    ws[box_code_to_sheet_code(box_code)].font = fontans
    ws[box_code_to_sheet_code(box_code)] = value
    wb.save("Data.xlsx")


def box_code_to_sheet_code(box_code):
    return box_code[1].upper() + str(ord(box_code[0]) - 54)

def check_bl_num(box_code):
    if box_code <= 'Ac' or 'Ba' <= box_code <= 'Bc' or 'Ca' <= box_code <= 'Cc':
        bl_num = 1
    elif 'Ad' <= box_code <= 'Af' or 'Bd' <= box_code <= 'Bf' or 'Cd' <= box_code <= 'Cf':
        bl_num = 2
    elif 'Ag' <= box_code <= 'Ai' or 'Bg' <= box_code <= 'Bi' or 'Cg' <= box_code <= 'Ci':
        bl_num = 3
    elif 'Da' <= box_code <= 'Dc' or 'Ea' <= box_code <= 'Ec' or 'Fa' <= box_code <= 'Fc':
        bl_num = 4
    elif 'Dd' <= box_code <= 'Df' or 'Ed' <= box_code <= 'Ef' or 'Fd' <= box_code <= 'Ff':
        bl_num = 5
    elif 'Dg' <= box_code <= 'Di' or 'Eg' <= box_code <= 'Ei' or 'Fg' <= box_code <= 'Fi':
        bl_num = 6
    elif 'Ga' <= box_code <= 'Gc' or 'Ha' <= box_code <= 'Hc' or 'Ia' <= box_code <= 'Ic':
        bl_num = 7
    elif 'Gd' <= box_code <= 'Gf' or 'Hd' <= box_code <= 'Hf' or 'Id' <= box_code <= 'If':
        bl_num = 8
    elif 'Gg' <= box_code <= 'Gi' or 'Hg' <= box_code <= 'Hi' or 'Ig' <= box_code <= 'Ii':
        bl_num = 9

    return bl_num

print('Processing..')
wb = load_workbook('Data.xlsx')
ws = wb.active

data = dict({})

# Assigning master dictionary with all numbers
for i in range(1, 10):

    for j in range(65, 74):
        k = chr(j)
        val = ws[k + str(i)].value
        key = str(chr(i + 64)) + str(chr(j + 32))
        data[key] = val
Qdata = {}
for j in data:
    if data[j] is not None:
        Qdata[j] = data[j]

# Assigning datas to individual 9-cell blocks
for i in range(1,10):
    globals()['bl'+str(i)] = {}
for i in data:
    if i <= 'Ac' or 'Ba' <= i <= 'Bc' or 'Ca' <= i <= 'Cc':
        bl1[i] = (data[i])

for i in data:
    if 'Ad' <= i <= 'Af' or 'Bd' <= i <= 'Bf' or 'Cd' <= i <= 'Cf':
        bl2[i] = (data[i])

for i in data:
    if 'Ag' <= i <= 'Ai' or 'Bg' <= i <= 'Bi' or 'Cg' <= i <= 'Ci':
        bl3[i] = (data[i])

for i in data:
    if 'Da' <= i <= 'Dc' or 'Ea' <= i <= 'Ec' or 'Fa' <= i <= 'Fc':
        bl4[i] = (data[i])

for i in data:
    if 'Dd' <= i <= 'Df' or 'Ed' <= i <= 'Ef' or 'Fd' <= i <= 'Ff':
        bl5[i] = (data[i])

for i in data:
    if 'Dg' <= i <= 'Di' or 'Eg' <= i <= 'Ei' or 'Fg' <= i <= 'Fi':
        bl6[i] = (data[i])

for i in data:
    if 'Ga' <= i <= 'Gc' or 'Ha' <= i <= 'Hc' or 'Ia' <= i <= 'Ic':
        bl7[i] = (data[i])

for i in data:
    if 'Gd' <= i <= 'Gf' or 'Hd' <= i <= 'Hf' or 'Id' <= i <= 'If':
        bl8[i] = (data[i])

for i in data:
    if 'Gg' <= i <= 'Gi' or 'Hg' <= i <= 'Hi' or 'Ig' <= i <= 'Ii':
        bl9[i] = (data[i])

# Assigning data to rows
for i in range(65, 74):
    list(globals()).append(chr(i))
    globals()[chr(i)] = {}

for i in range(65, 74):
    for j in data:
        if j[0] == chr(i):
            globals()[chr(i)][j] = data[j]

# Assigning data to coloumns
for j in range(97, 106):
    list(globals()).append(chr(j))
    globals()[chr(j)] = {}

for k in range(97, 106):
    for j in data:
        if j[1] == chr(k):
            globals()[chr(k)][j] = data[j]

fillNum = {}
no_empt_box = 1
for j in data:
    if data[j] is None:
        globals()['fill' + str(no_empt_box)] = [j, None]
        fillNum['fill' + str(no_empt_box)] = [j, None]
        no_empt_box += 1
no_empt_box -= 1
num = 0
while num < len(fillNum):

    num += 1
    prev_num = False
    row = globals()['fill' + str(num)][0][0]
    col = globals()['fill' + str(num)][0][1]
    blvar = 'bl'+str(check_bl_num(globals()['fill' + str(num)][0]))
    if globals()['fill' + str(num)][1] is None:
        start_num = 1
    else:
        start_num = (globals()['fill' + str(num)][1]) + 1

    if start_num == 10:
        globals()['fill' + str(num)][1] = None
        assign_values_to_vars(globals()['fill' + str(num)][0], None)
        num -= 2
        continue

    count = start_num - 1
    for j in range(start_num, 10):
        if j in globals()[row].values() or j in globals()[col].values() or j in globals()[blvar].values():
            count += 1
            if count == 9:
                prev_num = True
                break
            continue

        else:
            globals()['fill' + str(num)][1] = j
            fillNum['fill' + str(num)][1] = j
            assign_values_to_vars(globals()['fill' + str(num)][0], j)
            break
    if prev_num:
        globals()['fill' + str(num)][1] = None
        assign_values_to_vars(globals()['fill' + str(num)][0], None)
        num -= 2

fontq = Font(bold=True,
             color='000000',
             size=14,
             name='Calibri')
fontans = Font(bold=False,
               color='0000FF',
               size=14,
               name='Calibri')

for k in Qdata:
    write_qnum_in_sheet(k, data[k])

for k in fillNum:
    write_num_in_sheet(fillNum[k][0], fillNum[k][1])

print('Done! Check the excel file for solution')
time.sleep(3)
